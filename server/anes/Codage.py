import os
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# Fonction pour afficher les colonnes du fichier Excel
def afficher_colonnes(fichier_excel):
    try:
        df = pd.read_excel(fichier_excel)
        colonnes = df.columns.tolist()
        print(f"Les colonnes du fichier sont : {colonnes}")
        return colonnes
    except FileNotFoundError:
        print(f"Le fichier '{fichier_excel}' n'a pas été trouvé.")
    except Exception as e:
        print(f"Erreur lors du chargement du fichier : {str(e)}")

# Fonction pour obtenir les valeurs d'une colonne sous forme de liste
def recuperer_colonne(fichier_excel, nom_colonne):
    try:
        df = pd.read_excel(fichier_excel)
        if nom_colonne in df.columns:
            return df[nom_colonne].tolist()
        else:
            print(f"La colonne '{nom_colonne}' n'existe pas dans le fichier.")
    except Exception as e:
        print(f"Erreur lors du chargement du fichier : {str(e)}")

# Fonction pour obtenir les valeurs uniques d'une colonne et les encoder
def coder_valeurs(valeurs_uniques):
    codage = {valeur: f'{bin(2**i)[2:]:0>{len(valeurs_uniques)}}' for i, valeur in enumerate(valeurs_uniques)}
    print(f"Codage binaire des valeurs uniques : {codage}")
    return codage

# Fonction pour appliquer le codage binaire sous forme de matrice de bits
def appliquer_codage_en_matrice(valeurs, codage):
    matrice_bits = []
    for val in valeurs:
        if val in codage:
            bits = [int(bit) for bit in codage[val]]
            matrice_bits.append(bits)
    matrice = np.array(matrice_bits)
    print("Matrice de codage binaire :")
    print(matrice)
    return matrice

# Fonction pour calculer la matrice de dissemblance entre les lignes
def calculer_distance_dissemblance(matrice_bits):
    
    matrice_bits = np.array(matrice_bits)

    
    matrice_bits = matrice_bits[1:, 1:] 
    
    matrice_bits = np.nan_to_num(matrice_bits, nan=-1) 
    n = matrice_bits.shape[0] 
    matrice_distance = np.zeros((n, n))  
    
    
    for i in range(n):
        for j in range(n):
            dissemblance = np.sum(matrice_bits[i] != matrice_bits[j]) / matrice_bits.shape[1]
            matrice_distance[i, j] = dissemblance

    # Print the dissimilarity matrix
    print("Matrice de dissemblance dans fonction:")
    print(matrice_distance)
    
    
    
    return matrice_distance


# def sauvegarder_matrices_dans_fichier_unique(matrice_codage, nom_colonne, fichier_sortie):
#     # Convertir matrice_codage en un tableau NumPy si ce n'est pas déjà le cas
#     matrice_codage = np.array(matrice_codage)

#     # Handle NaN values by replacing them with a specific value or dropping rows/columns as needed
#     matrice_codage = np.nan_to_num(matrice_codage, nan=-1)  # Remplacer NaN avant d'écrire

#     # Vérifie si le fichier existe déjà
#     try:
#         wb = openpyxl.load_workbook(fichier_sortie)
#         ws = wb.active
#     except FileNotFoundError:
#         wb = openpyxl.Workbook()
#         ws = wb.active
#         # Si le fichier n'existe pas, crée une première ligne vide
#         ws.append([])

#     # Vérifie si le nom de la colonne principale existe déjà
#     if nom_colonne not in [cell.value for cell in ws[1]]:
#         start_col = ws.max_column + 1  # Trouve la colonne de départ
#         end_col = start_col + matrice_codage.shape[1] - 1
#         ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
#         ws.cell(row=1, column=start_col).value = nom_colonne
#         ws.cell(row=1, column=start_col).alignment = Alignment(horizontal='center')

#         # Ajoute les sous-colonnes avec un nom unique basé sur nom_colonne
#         for i in range(matrice_codage.shape[1]):
#             ws.cell(row=2, column=start_col + i, value=f"{nom_colonne}_{i + 1}")

#     # Remplit les données de la matrice de codage sous les colonnes correspondantes
#     for i, row in enumerate(matrice_codage, start=3):  # Débute à la 3ème ligne pour les données
#         for j, val in enumerate(row):
#             ws.cell(row=i, column=start_col + j, value=val)

#     # Sauvegarde le fichier
#     wb.save(fichier_sortie)
#     print(f"Les données pour '{nom_colonne}' ont été sauvegardées avec succès dans '{fichier_sortie}'.")

    
def coder_ordinal(liste_ordoner):
    nombre_de_valeurs = len(liste_ordoner)
    matrice = [[0 for _ in range(nombre_de_valeurs)] for _ in range(nombre_de_valeurs)]

    for i in range(nombre_de_valeurs):
        for j in range(i, nombre_de_valeurs):
            matrice[i][j] = 1
    return matrice


def tablec_ordinal(matrice, colonne_liste, liste_ordoner):
    colomn_coder = []
    for element in colonne_liste:
        if element in liste_ordoner:
            index = liste_ordoner.index(element)
            ligne_specifique = matrice[index]
            colomn_coder.append(ligne_specifique)
        else:
            print(f"Warning: '{element}' not found in liste_ordoner.")
    return colomn_coder

def trier_selon_ordre_automatique(valeurs):
    # Afficher les valeurs uniques et demander à l'utilisateur de spécifier l'ordre
    print("Voici les valeurs uniques dans la colonne :")
    print(valeurs)

    ordre_personnalise = {}
    
    # Demander à l'utilisateur d'entrer l'ordre souhaité pour chaque valeur
    for valeur in valeurs:
        position = input(f"Entrez la position pour la valeur '{valeur}' (entier positif) : ")
        ordre_personnalise[valeur] = int(position)
    
    # Trier les valeurs selon les positions données par l'utilisateur
    valeurs_triees = sorted(ordre_personnalise, key=ordre_personnalise.get)
    
    return valeurs_triees
def calculer_burts(matrice_codage):
    print("raaniiiiiiii f burttttt")
    matrice_codage = np.nan_to_num(np.array(matrice_codage), nan=-1)  # Replace NaN with -1
    nb_colon = matrice_codage.shape[1]
    matrice_burt = np.zeros((nb_colon, nb_colon))

    for i in range(nb_colon):
        for j in range(nb_colon):
            matrice_burt[i][j] = np.sum(matrice_codage[:, i] * matrice_codage[:, j])
            
            # Co-occurrence calculation
    
    
    print(f"rani ngl3 column talya  {matrice_burt.shape[0]}") # Remove last row and last column
    print(matrice_burt)
    return matrice_burt




def clc_table_de_conti(dictionnaire, matriceburt, variable1, variable2):
    # Convert matriceburt to a NumPy array
    matriceburt = np.array(matriceburt)

    
    
    # Retrieve start and end indexes based on variable names in the dictionary
    try:
        col_1, fin1 = dictionnaire[variable1]
        col_2, fin2 = dictionnaire[variable2]
        
    
    except KeyError as e:
        print(f"Erreur: la variable {e} n'existe pas dans le dictionnaire.")
        return
    
    # Initialize the contingency table with correct size
    contingency_table = np.zeros((fin1 - col_1 + 1, fin2 - col_2 + 1))
    indiceligne = 0
    
    # Iterate over the specified ranges for rows and columns
    for i in range(col_1, min(fin1 + 1, matriceburt.shape[0])):  # Prevent out of bounds on the row
        indicecolumn = 0
        for j in range(col_2, min(fin2 + 1, matriceburt.shape[0])):  # Prevent out of bounds on the column
            
            contingency_table[indiceligne][indicecolumn] = matriceburt[i][j]
            indicecolumn += 1
        indiceligne += 1

    print("Contingency Table Computed.")
    
    # Uncomment this block if saving to Excel is required
    # df_contingency = pd.DataFrame(contingency_table)
    # df_contingency.to_excel('contingency_table.xlsx', index=False, header=False)
    # print(f"Table de contingence sauvegardée dans le fichier: contingency_table.xlsx")
    
    print(contingency_table)
    return contingency_table


def calc_frequencies(matrice_frec):
    print("Calculating frequencies...")

    # Calculate the sum of all elements in the matrix
    somme = np.sum(matrice_frec)
    
    # Normalize the matrix such that the sum is 1
    if somme != 0:  # Avoid division by zero
        matrice_frec = matrice_frec / somme  # Normalize the matrix
    else:
        print("The sum of elements is zero, normalization cannot be performed.")
        return None  # Return None if normalization cannot be done due to zero sum
    
    print("Normalized matrix:")
    print(matrice_frec)
    return matrice_frec  # Return the normalized matrix

def calc_profitligne(matrice_frec):
    print(" iam profitttt")
    
    matrice_proft = np.nan_to_num(np.array(matrice_frec), nan=-1) 
    nbligne = matrice_frec.shape[0]
    
    nb_colmn = matrice_frec.shape[1]
    
        
    # Diviser chaque élément par la somme de sa ligne
    for i in range(nbligne):
        somme_ligne = np.sum(matrice_frec[i]) 
        # Calcule la somme de la ligne i
        
        if somme_ligne != 0:  # Vérifie que la somme n'est pas zéro pour éviter la division par zéro
            matrice_proft[i] = matrice_proft[i] / somme_ligne
        else:
                print(f"Somme de zéro évitée à la ligne {i}.")
        
        
        
        print(matrice_proft)
        print("donnne with profitligne")
    return matrice_proft  # Retourner la matrice modifiée si besoin



import numpy as np

def calculer_nuage_points(matrice_proft, matrmatrice_frec):
    nuage = {}
    
    # Obtenir le nombre de lignes dans matrice_proft ou matrmatrice_frec
    nbligne = len(matrice_proft)
    print(f"Nombre de lignes: {nbligne}")  # Debugging
    
    for i in range(nbligne):
        # Calcule la somme de la ligne i dans matrmatrice_frec
        somme_ligne = np.sum(matrmatrice_frec[i])
        print(f"Somme de la ligne {i} dans matrmatrice_frec: {somme_ligne}")  # Debugging
        
        # Récupère la ligne correspondante de matrice_proft
        ligne = matrice_proft[i]
        print(f"Ligne {i} dans matrice_proft: {ligne}")  # Debugging
        
        # Ajoute l'entrée au dictionnaire nuage
        nuage[i] = (ligne, somme_ligne)
    
    print(f"Nuage de points: {nuage}")  # Debugging
    return nuage

def calculer_centre_gravite(nuage):
    print(nuage)
    print("calculer_centre_gravite")
    # Initialiser la somme des coordonnées et la somme des fréquences
    somme_points = np.zeros(len(next(iter(nuage.values()))[0]))  # Taille de chaque ligne (dimension du point)
    somme_frequences = 0
    print(f"Somme initiale des points: {somme_points}, Somme des fréquences: {somme_frequences}")  # Debugging
    
    for i in nuage:
        ligne, frequence = nuage[i]
        
        # Ajouter la ligne multipliée par la fréquence à la somme des points
        somme_points += np.array(ligne) * frequence
        somme_frequences += frequence
        
        print(f"Après traitement du point {i}: somme_points = {somme_points}, somme_frequences = {somme_frequences}")  # Debugging
    
    # Vérifier que la somme des fréquences n'est pas nulle pour éviter la division par zéro
    if somme_frequences == 0:
        raise ValueError("La somme des fréquences est égale à zéro, le centre de gravité ne peut pas être calculé.")
    
    # Calculer le centre de gravité (moyenne pondérée par les fréquences)
    centre_gravite = somme_points / somme_frequences
    print(f"Centre de gravité calculé: {centre_gravite}")  # Debugging
    
    return centre_gravite

def calculer_inertie(nuage, centre_gravite):

    # Initialiser l'inertie
    inertie = 0
    print(f"Centre de gravité pour calculer l'inertie: {centre_gravite}")  # Debugging
    
    for i in nuage:
        ligne, frequence = nuage[i]
        
        # Calculer la distance au carré entre le point et le centre de gravité
        distance_carre = np.sum((np.array(ligne) - centre_gravite) ** 2)
        print(f"Point {i}: ligne = {ligne}, centre_gravite = {centre_gravite}, distance_carre = {distance_carre}")  # Debugging
        
        # Ajouter la contribution pondérée par la fréquence à l'inertie
        inertie += frequence * distance_carre
        print(f"Inertie après traitement du point {i}: {inertie}")  # Debugging
    
    return inertie
def convert_ndarray(data):
    if isinstance(data, dict):
        return {k: convert_ndarray(v) for k, v in data.items()}
    elif isinstance(data, list):
        return [convert_ndarray(v) for v in data]
    elif isinstance(data, np.ndarray):
        return data.tolist()